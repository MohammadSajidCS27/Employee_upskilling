from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


class WorkbookSkillRepository:
    """Workbook-driven skill source using the team skill matrix XLSX file."""

    SHEET_ROLE_HINTS: Dict[str, List[str]] = {
        "BE Developer Skills": ["backend", "be developer", "java developer", "microservice", "spring", "java", "spring boot", "python backend", "api", "software developer", "software engineer"],
        "DevOps Engineer Skills": ["devops", "sre", "platform", "cloud engineer", "infrastructure", "docker", "kubernetes", "jenkins", "terraform", "ci/cd"],
        "UI Developers Skills": ["ui", "frontend", "front end", "react", "angular", "javascript", "typescript", "html", "css", "vue", "software developer", "software engineer"],
        "Design and Architecture Skills": ["architect", "architecture", "design pattern", "solution architect", "system design", "software developer", "software engineer"],
        "DB Related Skills": ["database", "db", "sql", "dba", "postgres", "mysql", "oracle", "mongodb", "redis", "data engineer", "software developer", "software engineer"],
        "Test Professional Skills": ["test", "qa", "sdet", "quality assurance", "automation testing", "selenium", "junit", "testng", "pytest", "cypress", "jest", "mocha", "postman", "manual testing", "functional testing", "regression testing", "system testing", "integration testing", "api testing", "test case", "bug tracking", "smoke testing", "sanity testing", "test execution", "end-to-end testing", "cross-browser testing", "database validation", "istqb", "agile tester", "bug", "defect", "quality"],
        "PTE Skills": ["pte", "performance test", "performance engineer", "jmeter", "load test"],
        "UX Designer Skills": ["ux", "ui/ux", "product designer", "interaction designer", "figma"],
        "Scrum Master skills": ["scrum", "agile coach", "scrum master", "agile"],
    }

    def __init__(self, workbook_path: str):
        self.workbook_path = Path(workbook_path)
        self._sheet_skill_rows: Dict[str, List[Dict[str, Any]]] = {}
        self._all_skills: List[str] = []
        self._all_role_names: List[str] = []
        self._loaded = False

    @staticmethod
    def _normalize(value: str) -> str:
        return " ".join(str(value or "").strip().lower().split())

    @staticmethod
    def _parse_level(text: Any) -> int:
        if text is None:
            return 0
        raw = str(text).strip()
        if not raw:
            return 0
        numbers = [int(match) for match in re.findall(r"\d+", raw)]
        if not numbers:
            return 0
        return max(numbers)

    def _ensure_loaded(self) -> None:
        if self._loaded:
            return
        if not self.workbook_path.exists():
            logger.warning("Workbook not found: %s", self.workbook_path)
            self._loaded = True
            return

        workbook = load_workbook(str(self.workbook_path), data_only=True)

        all_skills: List[str] = []
        role_names: List[str] = []
        sheet_rows: Dict[str, List[Dict[str, Any]]] = {}

        for sheet_name in self.SHEET_ROLE_HINTS.keys():
            if sheet_name not in workbook.sheetnames:
                continue
            sheet = workbook[sheet_name]
            parsed = self._parse_role_sheet(sheet)
            if not parsed:
                continue
            sheet_rows[sheet_name] = parsed
            all_skills.extend([row["skill"] for row in parsed if row.get("skill")])

        for sheet_name, hints in self.SHEET_ROLE_HINTS.items():
            role_names.append(self._normalize(sheet_name.replace(" skills", "").replace(" Skills", "")))
            role_names.extend([self._normalize(hint) for hint in hints])

        self._sheet_skill_rows = sheet_rows
        self._all_skills = sorted(set(skill for skill in all_skills if skill))
        self._all_role_names = sorted(set(role for role in role_names if role))
        self._loaded = True

        logger.info(
            "Workbook skill repository loaded from '%s' with sheets=%s skills=%s",
            self.workbook_path,
            len(self._sheet_skill_rows),
            len(self._all_skills),
        )

    def _locate_header(self, sheet: Any) -> Optional[Tuple[int, int]]:
        for row in range(1, min(sheet.max_row, 20) + 1):
            for col in range(1, min(sheet.max_column, 12) + 1):
                value = sheet.cell(row=row, column=col).value
                if self._normalize(value) in {"skill area", "skill"}:
                    return row, col
        return None

    def _parse_role_sheet(self, sheet: Any) -> List[Dict[str, Any]]:
        header = self._locate_header(sheet)
        if not header:
            return []

        header_row, skill_col = header
        junior_col = skill_col + 1
        middle_col = skill_col + 2
        senior_col = skill_col + 3

        output: List[Dict[str, Any]] = []
        current_heading: Optional[str] = None
        headings_with_skills: set = set()
        
        for row in range(header_row + 1, sheet.max_row + 1):
            raw_skill = sheet.cell(row=row, column=skill_col).value
            skill = self._normalize(raw_skill)
            if not skill:
                continue
            if skill in {"skill area", "actual level", "training topics"}:
                continue
            if skill.startswith("skill matrix for"):
                continue

            cell = sheet.cell(row=row, column=skill_col)
            is_bold = bool(cell.font and cell.font.bold)
            
            if is_bold:
                current_heading = skill
                continue

            junior_level = self._parse_level(sheet.cell(row=row, column=junior_col).value)
            middle_level = self._parse_level(sheet.cell(row=row, column=middle_col).value)
            senior_level = self._parse_level(sheet.cell(row=row, column=senior_col).value)
            if max(junior_level, middle_level, senior_level) == 0:
                continue

            headings_with_skills.add(current_heading)
            output.append(
                {
                    "skill": skill,
                    "heading": current_heading,
                    "levels": {
                        "junior": junior_level,
                        "middle": middle_level,
                        "senior": senior_level,
                    },
                }
            )

        # Add standalone bold items (headings with no sub-skills) as regular skills
        for row in range(header_row + 1, sheet.max_row + 1):
            raw_skill = sheet.cell(row=row, column=skill_col).value
            skill = self._normalize(raw_skill)
            if not skill or skill in {"skill area", "actual level", "training topics"}:
                continue
            if skill.startswith("skill matrix for"):
                continue
            cell = sheet.cell(row=row, column=skill_col)
            is_bold = bool(cell.font and cell.font.bold)
            if is_bold and skill not in headings_with_skills:
                output.append(
                    {
                        "skill": skill,
                        "heading": None,
                        "levels": {
                            "junior": self._parse_level(sheet.cell(row=row, column=junior_col).value),
                            "middle": self._parse_level(sheet.cell(row=row, column=middle_col).value),
                            "senior": self._parse_level(sheet.cell(row=row, column=senior_col).value),
                        },
                    }
                )

        deduped: Dict[str, Dict[str, Any]] = {}
        for item in output:
            key = item["skill"]
            if key not in deduped:
                deduped[key] = item
                continue
            existing = deduped[key]
            for bucket in ("junior", "middle", "senior"):
                existing["levels"][bucket] = max(existing["levels"][bucket], item["levels"][bucket])
        return list(deduped.values())

    def get_all_skills(self) -> List[str]:
        self._ensure_loaded()
        return list(self._all_skills)

    def get_role_keywords(self) -> List[str]:
        self._ensure_loaded()
        return list(self._all_role_names)

    def _score_sheets(self, search_text: str) -> Dict[str, int]:
        """Score every sheet by how many of its role hints appear in ``search_text``.

        Multi-word hints score higher than generic single tokens so specific role
        phrases (e.g. "java developer") outweigh broad ones (e.g. "developer").
        """
        scores: Dict[str, int] = {}
        for sheet_name, hints in self.SHEET_ROLE_HINTS.items():
            score = 0
            for hint in hints:
                normalized_hint = self._normalize(hint)
                if normalized_hint and normalized_hint in search_text:
                    score += len(normalized_hint.split())
            scores[sheet_name] = score
        return scores

    def _resolve_sheet_for_role(self, role: str, current_skills: Optional[List[str]] = None) -> Optional[str]:
        normalized_role = self._normalize(role)
        normalized_skills = " ".join(self._normalize(skill) for skill in (current_skills or []))
        sheet_order = list(self.SHEET_ROLE_HINTS.keys())

        # The TARGET ROLE is the user's intent (where they want to go) and must drive the
        # expected-skill set. Current skills describe where the candidate is today, so they
        # must NOT override the role when selecting the sheet -- otherwise every target role
        # collapses onto whichever sheet the resume skills happen to match.
        if normalized_role:
            role_scores = self._score_sheets(normalized_role)
            top_score = max(role_scores.values(), default=0)
            if top_score > 0:
                leaders = [name for name in sheet_order if role_scores[name] == top_score]
                if len(leaders) == 1:
                    return leaders[0]
                # Ambiguous role (e.g. a generic "software engineer" title matches several
                # sheets): break the tie using current-skill evidence, but only among the
                # sheets the role already qualifies for.
                skill_scores = self._score_sheets(normalized_skills)
                leaders.sort(key=lambda name: (-skill_scores.get(name, 0), sheet_order.index(name)))
                return leaders[0]

            # Generic software/engineering titles with no specific sheet default to backend.
            if any(token in normalized_role for token in ["software", "developer", "engineer"]):
                return "BE Developer Skills"
            # A role with no workbook coverage (e.g. "Data Scientist") stays unmapped so the
            # caller can fall back to ESCO/O*NET/LLM sources.
            return None

        # No role signal at all: fall back to skills evidence only.
        if normalized_skills:
            skill_scores = self._score_sheets(normalized_skills)
            top_score = max(skill_scores.values(), default=0)
            if top_score > 0:
                return next(name for name in sheet_order if skill_scores[name] == top_score)
        return None

    @staticmethod
    def _experience_bucket(experience_years: int) -> str:
        if experience_years < 4:
            return "junior"
        if experience_years < 8:
            return "middle"
        return "senior"

    def get_all_skills_for_role(self, role: str, experience_years: int = 0, current_skills: Optional[List[str]] = None) -> Dict[str, Any]:
        """Get all skills from the specific workbook sheet matched to the role.
        
        This method ensures ALL skills from the matched sheet are returned (not filtered by level),
        enabling complete skill gap analysis against the role's skill requirements.
        """
        self._ensure_loaded()
        
        sheet = self._resolve_sheet_for_role(role, current_skills=current_skills)
        if not sheet:
            return {
                "skills": [],
                "source": "workbook_role_unmapped",
                "live": False,
                "provider": "workbook",
                "sheet": None,
                "workbook_file": str(self.workbook_path),
            }
        
        rows = self._sheet_skill_rows.get(sheet, [])
        if not rows:
            return {
                "skills": [],
                "source": "workbook_sheet_empty",
                "live": False,
                "provider": "workbook",
                "sheet": sheet,
            }
        
        bucket = self._experience_bucket(int(experience_years or 0))
        
        # Return ALL skills from the sheet (not filtered by level threshold)
        all_skills = [row["skill"] for row in rows if row.get("skill")]
        
        skill_levels = {
            row["skill"]: int(row.get("levels", {}).get(bucket, 0))
            for row in rows
            if row.get("skill")
        }
        
        heading_map: Dict[str, List[str]] = {}
        for row in rows:
            heading = row.get("heading")
            if heading:
                heading_map.setdefault(heading, []).append(row["skill"])
        
        return {
            "skills": sorted(set(all_skills)),
            "source": f"workbook:{sheet}",
            "live": False,
            "provider": "workbook",
            "sheet": sheet,
            "workbook_file": str(self.workbook_path),
            "headings": heading_map,
            "skill_levels": skill_levels,
            "experience_bucket": bucket,
        }

    @staticmethod
    def _is_software_related_role(normalized_role: str, normalized_skills: str) -> bool:
        """Check if the role or skills indicate a software/technical role."""
        software_markers = {"software", "developer", "engineer", "backend", "frontend", "fullstack", 
                           "java", "python", "javascript", "typescript", "react", "angular", 
                           "docker", "kubernetes", "aws", "cloud", "api", "microservice"}
        combined = f"{normalized_role} {normalized_skills}"
        return any(marker in combined for marker in software_markers)

    def get_skill_lookup_result(self, role: str, experience_years: int = 0, current_skills: Optional[List[str]] = None) -> Dict[str, Any]:
        self._ensure_loaded()
        sheet = self._resolve_sheet_for_role(role, current_skills=current_skills)
        if not sheet:
            return {
                "skills": [],
                "source": "workbook_role_unmapped",
                "live": False,
                "provider": "workbook",
                "sheet": None,
                "workbook_file": str(self.workbook_path),
            }

        rows = self._sheet_skill_rows.get(sheet, [])
        if not rows:
            return {
                "skills": [],
                "source": "workbook_sheet_empty",
                "live": False,
                "provider": "workbook",
                "sheet": sheet,
            }

        bucket = self._experience_bucket(int(experience_years or 0))
        min_required = 2 if bucket in {"junior", "middle"} else 3

        selected = [
            row["skill"]
            for row in rows
            if int(row.get("levels", {}).get(bucket, 0)) >= min_required
        ]

        # If very sparse results in the selected bucket, fallback to all sheet skills.
        if len(selected) < 20:
            selected = [row["skill"] for row in rows]

        skill_levels = {
            row["skill"]: int(row.get("levels", {}).get(bucket, 0))
            for row in rows
            if row.get("skill")
        }

        heading_map: Dict[str, List[str]] = {}
        for row in rows:
            heading = row.get("heading")
            if heading:
                heading_map.setdefault(heading, []).append(row["skill"])

        return {
            "skills": sorted(set(selected)),
            "source": f"workbook:{sheet}",
            "live": False,
            "provider": "workbook",
            "sheet": sheet,
            "workbook_file": str(self.workbook_path),
            "headings": heading_map,
        }
